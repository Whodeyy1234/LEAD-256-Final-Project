from tkinter import *
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter import messagebox as mb
from tkinter import simpledialog as sd
from itertools import chain
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import *

class App(Tk):
    def __init__(self):
        self.path = ''

        self.changeFile()
            
        super().__init__()
        
        ttk.Label(self, text = 'Main Menu').pack()

        showBtn = ttk.Button(self, text = 'Show Data', command = self.showData).pack()
        createBtn = ttk.Button(self, text = 'Create Data', command = self.createData).pack()
        insertBtn = ttk.Button(self, text = 'Insert Data', command = lambda: self.manipulateData('insert')).pack()
        deleteBtn = ttk.Button(self, text = 'Delete Data', command = lambda: self.manipulateData('delete')).pack()
        macroBtn = ttk.Button(self, text = 'Macros', command = self.macroMenu).pack()
        fileBtn = ttk.Button(self, text = 'Change File', command = self.changeFile).pack()
        exitBtn = ttk.Button(self, text = 'Exit', command = self.destroy).pack()

        self.geometry('{}x{}'.format(175,200))
        self.centerWindow(self)

    def showData(self):
        self.destroyChildren()
        showForm = Toplevel(self)

        df = pd.read_excel(self.path)
        try:
            cols = len(df.columns)

            headings, data = self.dictionaryToData(df)

            for i in range(len(headings)):
                ttk.Label(showForm, text = headings[i]).grid(column = i + 1, row = 0)
            
            for i in range(len(data[i])):
                ttk.Label(showForm, text = str(i + 1)).grid(column = 0, row = i + 1)
            
            for i in range(len(data)):
                for j in range(len(data[i])):
                    ttk.Label(showForm, text = data[i][j],
                              backgroun = 'white', foreground = 'black',
                              width = 10, anchor = 'n').grid(column = i + 1, row = j + 1)
                    showForm.update()
            
        except:
            ttk.Label(showForm, text = 'No data to display!').pack()

        showForm.update()
        self.centerWindow(showForm)

    def createData(self):
        self.destroyChildren()
        createForm = Toplevel(self)

        ttk.Label(createForm, text = '# of rows:').grid(column = 2, row = 0)
        ttk.Label(createForm, text = '# of columns/headings:').grid(column = 0, columnspan = 3, row = 1)

        numRows = Text(createForm, height = 1, width = 10)
        numRows.bind('<Tab>', self.focusNextWidget)
        numRows.grid(column = 3, row = 0)

        numCols = Text(createForm, height = 1, width = 10)
        numCols.bind('<Tab>', self.focusNextWidget)
        numCols.bind('<Return>', lambda x=None: self.submitToData(int(numRows.get('1.0','end')),
                                                                  int(numCols.get('1.0','end'))))
        numCols.grid(column = 3, row = 1)
        

        submitBtn = ttk.Button(createForm, text = 'Submit',
                               command = lambda : self.submitToData(int(numRows.get('1.0','end')),
                                                                   int(numCols.get('1.0','end'))))
        submitBtn.grid(column = 0,row = 2)

        clearBtn = ttk.Button(createForm, text = 'Clear', command = lambda: self.clearText([numRows,numCols]))
        clearBtn.grid(column = 3, row = 2)

        self.centerWindow(createForm)

    def manipulateData(self,name=''):
        self.destroyChildren()
        manuForm = Toplevel(self)

        ttk.Label(manuForm, text = 'Row # to ' + name + ' data:').grid(column = 0, row = 0)
        insertText = Text(manuForm, height = 1, width = 10)
        insertText.bind('<Tab>', self.focusNextWidget)
        insertText.grid(column = 1, row = 0)

        valueText = Text(manuForm, height = 1, width = 10)
        valueText.bind('<Tab>', self.focusNextWidget)
        valueText.grid(column = 1, row = 1)

        if name == 'insert':
            ttk.Label(manuForm, text = '# of Rows to insert:').grid(column = 0, row = 1)
            submitBtn = ttk.Button(manuForm, text = 'Submit',
                                   command = lambda: self.submitToInsert(int(insertText.get('1.0','end').rstrip()),
                                                                         int(valueText.get('1.0','end').rstrip())))
        else:
            ttk.Label(manuForm, text = '# of Rows to delete:').grid(column = 0, row = 1)
            submitBtn = ttk.Button(manuForm, text = 'Submit',
                                   command = lambda: self.submitToDelete(int(insertText.get('1.0','end').rstrip()),
                                                                         int(valueText.get('1.0','end').rstrip())))

        submitBtn.grid(column = 0, row = 2)

        clearBtn = ttk.Button(manuForm, text = 'Clear', command = lambda: self.clearText([insertText, valueText]))
        clearBtn.grid(column = 1, row = 2)

        self.centerWindow(manuForm)

    def macroMenu(self):
        self.destroyChildren()
        macroForm = Toplevel(self)
        self.path.replace('\\','/')

        ttk.Label(macroForm, text = 'Macro Menu').pack()

        tableBtn = ttk.Button(macroForm, text = 'Create Table', command = self.createTable).pack()
        sortAscendBtn = ttk.Button(macroForm, text = 'Sort Ascending', command = self.sortAscending).pack()
        sortDescendBtn = ttk.Button(macroForm, text = 'Sort Descending', command = self.sortDescending).pack()
        exitBtn = ttk.Button(macroForm, text = 'Exit', command = macroForm.destroy).pack()

        self.centerWindow(macroForm)

    def createTable(self):
        wb = load_workbook(self.path)
        ws = wb['Sheet1']
        table = Table(displayName = 'Table1', ref = 'A1:' +
                      get_column_letter(ws.max_column) + str(ws.max_row))
        ws.add_table(table)
        wb.save(self.path)

    def sortAscending(self):
        df = pd.read_excel(self.path)
        self.columnChooser(df,0)

    def sortDescending(self):
        df = pd.read_excel(self.path)
        self.columnChooser(df,1)

    def columnChooser(self,df,sort):
        chooserForm = Toplevel()
        ttk.Label(chooserForm, text = 'Choose which column to sort by:').pack()

        self.colChoice = StringVar()
        for i in range(len(df.columns)):
            ttk.Radiobutton(chooserForm, text = df.columns[i], variable = self.colChoice, value = df.columns[i]).pack()
        submitBtn = ttk.Button(chooserForm, text = 'Submit', command = lambda: [self.sortValues(df,sort),chooserForm.destroy()]).pack()

        self.centerWindow(chooserForm)

    def sortValues(self,df,sort):
        if(sort == 0):
            df.sort_values(ascending = True, by = self.colChoice.get(),inplace = True)
        elif(sort == 1):
            df.sort_values(ascending = False, by = self.colChoice.get(),inplace = True)
        self.writeToExcel(df)

    def submitToData(self, rows, cols):
        self.destroyChildren()
        dataForm = Toplevel(self)
        
        headingFields = []
        ttk.Label(dataForm, text = 'Headings:').grid(column = 0, row = 0)
        for j in range(cols):
            headingFields.append(Text(dataForm, height = 1, width = 10, name = 'heading ' + str(j)))
            headingFields[j].bind('<Tab>', self.focusNextWidget)
            headingFields[j].grid(column = j + 1, row = 0)

        wIds = []
        for i in range(rows*cols):
            wIds.append(i + 1)

        dataFields = []
        temp = []
        index = 0
        for i in range(rows):
            ttk.Label(dataForm, text = 'Row ' + str(i + 1) + ':').grid(column = 0, row = i + 1)
            for j in range(cols):
                temp.append(Text(dataForm, height = 1, width = 10, name = 'field ' + str(wIds[index])))
                index += 1
            dataFields.append(temp)
            temp = []

        for i in range(rows):
            for j in range(cols):
                dataFields[i][j].grid(column = j + 1, row = i + 1)
                dataFields[i][j].bind('<Tab>', self.focusNextWidget)

        submitBtn = ttk.Button(dataForm, text = 'Submit',
                               command = lambda: self.submitToExcel(headingFields,dataFields))
        submitBtn.grid(column = 0, row = rows + 1)

        clearBtn = ttk.Button(dataForm, text = 'Clear',
                              command = lambda: self.clearText(headingFields + list(chain.from_iterable(dataFields))))
        clearBtn.grid(column = cols, row = rows + 1)

        self.centerWindow(dataForm)

    def submitToExcel(self, headingFields, dataFields):
        headings = []
        for field in headingFields:
            headings.append(field.get('1.0','end').rstrip())
                            
        data = []
        temp = []
        for j in range(len(headings)):
            for i in range(len(dataFields)):
                try:
                    temp.append(int(dataFields[i][j].get('1.0','end').rstrip()))
                except:
                    temp.append(dataFields[i][j].get('1.0','end').rstrip())
            data.append(temp)
            temp = []

        dataDict = {}
        for key in headings:
            for value in data:
                dataDict[key] = value
                data.remove(value)
                break

        df = pd.DataFrame(dataDict)
        writer = pd.ExcelWriter(self.path, engine='xlsxwriter')
        df.to_excel(writer,index=False)
        writer.save()

        mb.showinfo('Data Submition', 'Your data has been saved!')
        
        self.destroyChildren()
        

    def submitToInsert(self,rowNum, rows):
        df = pd.read_excel(self.path)
        cols = len(df.columns)
        self.destroyChildren()
        dataForm = Toplevel(self)

        for i in range(len(df.columns)):
            ttk.Label(dataForm, text = df.columns[i]).grid(column = i + 1, row = 0)

        wIds = []
        for i in range(rows*cols):
            wIds.append(i + 1)

        dataFields = []
        temp = []
        index = 0
        for i in range(rows):
            ttk.Label(dataForm, text = 'Row ' + str(rowNum + i) + ':').grid(column = 0, row = i + 1)
            for j in range(cols):
                temp.append(Text(dataForm, height = 1, width = 10, name = 'field ' + str(wIds[index])))
                index += 1
            dataFields.append(temp)
            temp = []

        for i in range(rows):
            for j in range(cols):
                dataFields[i][j].grid(column = j + 1, row = i + 1)
                dataFields[i][j].bind('<Tab>', self.focusNextWidget)

        submitBtn = ttk.Button(dataForm, text = 'Submit',
                               command = lambda: self.insertToExcel(rowNum,rows,cols,dataFields,df))
        submitBtn.grid(column = 0, row = rows + 1)

        clearBtn = ttk.Button(dataForm, text = 'Clear',
                              command = lambda: self.clearText(list(chain.from_iterable(dataFields))))
        clearBtn.grid(column = cols, row = rows + 1)

        self.centerWindow(dataForm)

    def submitToDelete(self, rowNum, rows):
        df = pd.read_excel(self.path)
        cols = len(df.columns)

        headings, data = self.dictionaryToData(df)
        
        for i in range(len(data)):
            del data[i][rowNum - 1:rowNum+rows-1]

        dataDict = self.dataToDictionary(headings, data)

        self.writeToExcel(dataDict)
        
    def insertToExcel(self, numRows, rows, headings, dataFields, df):
        dataDict = df.to_dict()
        headings = []
        data = []
        temp = []
        for key in dataDict:
            headings.append(key)
            for value in dataDict[key]:
                temp.append(dataDict[key][value])
            data.append(temp)
            temp = []
        temp = []
        for j in range(len(headings)):
            for i in range(len(dataFields)):
                try:
                    temp.append(int(dataFields[i][j].get('1.0','end').rstrip()))
                except:
                    temp.append(dataFields[i][j].get('1.0','end').rstrip())
            for k in range(len(temp)):
                data[j].insert(numRows - 1,temp[rows-k-1])
            temp = []
        
        dataDict = {}
        tempDict = {}
        count = 0
        INDEX = 0
        for i in headings:
            for j in range(len(data[count])):
                tempDict[INDEX + j] = data[count][j]
            dataDict[i] = tempDict
            tempDict = {}
            count += 1

        self.writeToExcel(dataDict)

    def changeFile(self):
        root = Tk()

        ttk.Label(root, text = 'Please find an empty Excel file:').grid(column = 0, row = 0)

        filePathText = Text(root, height = 1, width = 50)
        filePathText.grid(column = 0, row = 1)
        browseBtn = ttk.Button(root, text = 'Browse...', command = lambda: self.getFilePath(filePathText))
        browseBtn.grid(column = 1, row = 1)
        beginBtn = ttk.Button(root, text = 'Begin', command = lambda: self.beginApp(filePathText, root))
        beginBtn.grid(column = 0, row = 2, sticky = 'W')
        createBtn = ttk.Button(root, text = 'Create', command = lambda: self.createFile(root))
        createBtn.grid(column = 1, row = 2)

        self.centerWindow(root)
        root.mainloop()

    def destroyChildren(self):
        for widget in self.winfo_children():
            if isinstance(widget, Toplevel):
                widget.destroy()

    def focusNextWidget(self, event):
        event.widget.tk_focusNext().focus()
        return('break')

    def clearText(self, fields):
        for i in fields:
            i.delete('1.0','end')

    def getFilePath(self, text):
        text.insert('1.0',fd.askopenfilename(filetypes = [('Excel files', '*.xlsx')]))

    def beginApp(self, text, root):
        if text.compare('end-1c','==','1.0'):
            mb.showerror('Python Error', 'Error: Please make sure to select a file!')
        else:
            self.path = text.get('1.0','end').rstrip()
            root.destroy()
        

    def createFile(self, root):
        mb.showinfo('Important Message', 'Please select a directory where you want the file to be made.')
        self.path = fd.askdirectory()
        name = sd.askstring('Input','Name your file:')
        self.path += '/' + name + '.xlsx'

        writer = pd.ExcelWriter(self.path, engine='xlsxwriter')
        writer.save()
        
        root.destroy()

    def centerWindow(self, win):
        win.resizable(False, False)
        win.update_idletasks()

        window_height = win.winfo_height()
        window_width = win.winfo_width()

        screen_width = win.winfo_screenwidth()
        screen_height = win.winfo_screenheight()

        x_cordinate = int((screen_width/2) - (window_width/2))
        y_cordinate = int((screen_height/2) - (window_height/2))

        win.geometry("{}x{}+{}+{}".format(window_width, window_height, x_cordinate, y_cordinate))

    def writeToExcel(self,data):
        if(type(data) is dict):
           df = pd.DataFrame(data)
        else:
            df = data
        writer = pd.ExcelWriter(self.path, engine='xlsxwriter')
        df.to_excel(writer,index=False)
        writer.save()

        mb.showinfo('Data Submition', 'Your data has been saved!')
        
        self.destroyChildren()

    def dictionaryToData(self, df):
        dataDict = df.to_dict()
        headings = []
        data = []
        temp = []
        for key in dataDict:
            headings.append(key)
            for value in dataDict[key]:
                temp.append(dataDict[key][value])
            data.append(temp)
            temp = []
        return headings, data

    def dataToDictionary(self, headings, data):
        dataDict = {}
        tempDict = {}
        count = 0
        INDEX = 0
        for i in headings:
            for j in range(len(data[count])):
                tempDict[INDEX + j] = data[count][j]
            dataDict[i] = tempDict
            tempDict = {}
            count += 1
        return dataDict
    
if __name__ == '__main__':
    app = App()
